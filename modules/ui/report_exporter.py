#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动报告导出 — 驾驶行为分析报告导出
═══════════════════════════════════════════════════════════════
支持格式:
  - Markdown: 纯文本 + 表格 + 统计摘要
  - Excel: 全指标表格 + 多Sheet (openpyxl)
  - JSON: 结构化数据导出
  - CSV: 事件列表 + 统计摘要
"""

import logging, json, os, time
from typing import Dict, List, Optional, Any
from datetime import datetime

logger = logging.getLogger(__name__)


class ReportExporter:
    """驾驶行为分析报告导出器

    用法:
        exporter = ReportExporter()
        exporter.export_markdown(events, stats, output_path)
        exporter.export_excel(events, stats, output_path)
        exporter.export_json(events, stats, output_path)
    """

    # 事件类型 -> 中文映射
    EVENT_CN_MAP = {
        'emergency_braking': '紧急制动',
        'aggressive_deceleration': '激进减速',
        'normal_deceleration': '正常减速',
        'aggressive_acceleration': '激进加速',
        'normal_acceleration': '正常加速',
        'launch': '起步',
        'constant_speed': '匀速',
        'stopped': '停车',
        'weaving': '蛇形驾驶',
        'lane_change': '变道',
        'rapid_direction_change': '急速变向',
        'tight_turn': '小半径转弯',
        'wide_turn': '大半径转弯',
        'u_turn': 'U型转弯',
        'straight_driving': '直线行驶',
        'lane_keeping': '车道保持',
        'cornering_acceleration': '弯道加速',
        'cornering_deceleration': '弯道减速',
        'cornering_braking': '弯道制动',
        'severe_bump': '剧烈颠簸',
        'skid_risk': '侧滑风险',
        'rollover_risk': '侧翻风险',
        'sensor_fault': '传感器异常',
        'normal': '正常驾驶',
    }

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def _cn(self, etype: str) -> str:
        return self.EVENT_CN_MAP.get(etype, etype)

    # ═══════════════════════════════════════════════════════════
    #  Markdown 导出
    # ═══════════════════════════════════════════════════════════

    def export_markdown(self, events: List[Dict], stats: Dict, output_path: str):
        """导出 Markdown 报告"""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        lines = [
            f"# 驾驶行为分析报告",
            f"",
            f"**生成时间**: {now}",
            f"**事件总数**: {stats.get('total_events', len(events))}",
            f"**分析时长**: {stats.get('duration', 'N/A')}",
            f"",
            f"---",
            f"",
            f"## 概览",
            f"",
            f"| 指标 | 值 |",
            f"|:---|---:|",
            f"| 总事件数 | {stats.get('total_events', 'N/A')} |",
            f"| 已确认 | {stats.get('confirmed', 'N/A')} |",
            f"| 待复核 | {stats.get('to_review', 'N/A')} |",
            f"| 误检 | {stats.get('rejected', 'N/A')} |",
            f"| 平均置信度 | {stats.get('avg_confidence', 'N/A')} |",
            f"",
            f"## 事件类型分布",
            f"",
            f"| 事件类型 | 数量 | 占比 |",
            f"|:---|---:|---:|",
        ]

        if stats.get('type_distribution'):
            for etype, count in sorted(stats['type_distribution'].items(), key=lambda x: x[1], reverse=True):
                pct = count / max(stats.get('total_events', 1), 1) * 100
                lines.append(f"| {self._cn(etype)} | {count} | {pct:.1f}% |")

        lines.extend([
            f"",
            f"## 事件详情",
            f"",
            f"| 时间 | 类型 | 置信度 | 来源 | 复核 |",
            f"|:---|---:|---:|:---|:---|",
        ])

        for evt in events[:200]:  # 最多200条
            ts = evt.get('timestamp', 0)
            etype = evt.get('event_type', evt.get('type', '??'))
            conf = evt.get('confidence', 0) * 100
            source = evt.get('source', '-')
            review = evt.get('review', '-')
            lines.append(f"| {ts:.1f}s | {self._cn(etype)} | {conf:.0f}% | {source} | {review} |")

        lines.extend([
            f"",
            f"---",
            f"",
            f"*报告由驾驶行为分析系统自动生成*",
        ])

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        self.logger.info(f"Markdown 报告已导出: {output_path}")
        return output_path

    # ═══════════════════════════════════════════════════════════
    #  Excel 导出
    # ═══════════════════════════════════════════════════════════

    def export_excel(self, events: List[Dict], stats: Dict, output_path: str):
        """导出 Excel 报告 (多Sheet)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.logger.error("openpyxl 未安装，无法导出 Excel")
            return None

        wb = openpyxl.Workbook()

        # Sheet 1: 概览
        ws = wb.active
        ws.title = "概览"

        header_font = Font(name='Microsoft YaHei', bold=True, size=12, color='2F5496')
        header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

        ws['A1'] = '驾驶行为分析报告'
        ws['A1'].font = Font(name='Microsoft YaHei', bold=True, size=16, color='2F5496')
        ws.merge_cells('A1:D1')

        ws['A3'] = '生成时间'
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A4'] = '事件总数'
        ws['B4'] = stats.get('total_events', len(events))
        ws['A5'] = '已确认'
        ws['B5'] = stats.get('confirmed', 0)
        ws['A6'] = '待复核'
        ws['B6'] = stats.get('to_review', 0)
        ws['A7'] = '误检'
        ws['B7'] = stats.get('rejected', 0)
        ws['A8'] = '平均置信度'
        ws['B8'] = f"{stats.get('avg_confidence', 0) * 100:.1f}%"

        # Sheet 2: 事件详情
        ws2 = wb.create_sheet("事件详情")
        headers = ['时间(s)', '事件类型', '置信度', '来源', '复核状态']
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, evt in enumerate(events[:1000], 2):
            ws2.cell(row=i, column=1, value=round(evt.get('timestamp', 0), 1))
            ws2.cell(row=i, column=2, value=self._cn(evt.get('event_type', evt.get('type', '??'))))
            ws2.cell(row=i, column=3, value=f"{evt.get('confidence', 0) * 100:.0f}%")
            ws2.cell(row=i, column=4, value=evt.get('source', '-'))
            ws2.cell(row=i, column=5, value=evt.get('review', '-'))

        # 调整列宽
        for ws_sheet in [ws, ws2]:
            for col in range(1, 6):
                ws_sheet.column_dimensions[get_column_letter(col)].width = 18

        # Sheet 3: 类型分布
        if stats.get('type_distribution'):
            ws3 = wb.create_sheet("事件类型分布")
            ws3.cell(row=1, column=1, value='事件类型').font = header_font
            ws3.cell(row=1, column=2, value='数量').font = header_font
            ws3.cell(row=1, column=3, value='占比').font = header_font
            ws3.cell(row=1, column=1).fill = header_fill
            ws3.cell(row=1, column=2).fill = header_fill
            ws3.cell(row=1, column=3).fill = header_fill

            for i, (etype, count) in enumerate(sorted(
                stats['type_distribution'].items(), key=lambda x: x[1], reverse=True
            ), 2):
                ws3.cell(row=i, column=1, value=self._cn(etype))
                ws3.cell(row=i, column=2, value=count)
                pct = count / max(stats.get('total_events', 1), 1) * 100
                ws3.cell(row=i, column=3, value=f"{pct:.1f}%")

        wb.save(output_path)
        self.logger.info(f"Excel 报告已导出: {output_path}")
        return output_path

    # ═══════════════════════════════════════════════════════════
    #  JSON 导出
    # ═══════════════════════════════════════════════════════════

    def export_json(self, events: List[Dict], stats: Dict, output_path: str):
        """导出 JSON 结构化数据"""
        report = {
            'generated_at': datetime.now().isoformat(),
            'summary': stats,
            'events': events,
        }
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2, default=str)

        self.logger.info(f"JSON 报告已导出: {output_path}")
        return output_path

    # ═══════════════════════════════════════════════════════════
    #  CSV 导出
    # ═══════════════════════════════════════════════════════════

    def export_csv(self, events: List[Dict], stats: Dict, output_path: str):
        """导出 CSV 文件"""
        import csv
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['timestamp', 'event_type', 'confidence', 'source', 'review'])
            for evt in events:
                writer.writerow([
                    evt.get('timestamp', 0),
                    evt.get('event_type', evt.get('type', '??')),
                    f"{evt.get('confidence', 0) * 100:.1f}%",
                    evt.get('source', '-'),
                    evt.get('review', '-'),
                ])

        self.logger.info(f"CSV 已导出: {output_path}")
        return output_path