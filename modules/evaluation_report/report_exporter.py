#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
诊断报告一键导出模块 (U5)

支持:
- PDF 导出 (WeasyPrint/ReportLab)
- Markdown 导出
- Excel 导出 (openpyxl)
- 图表嵌入
- 一键导出按钮
"""

import logging
import os
import json
from typing import Dict, List, Any, Optional
from datetime import datetime
from io import BytesIO

logger = logging.getLogger(__name__)


class ReportExporter:
    """诊断报告导出器 (U5)

    支持三种格式: PDF / Markdown / Excel
    """

    def __init__(self, output_dir: str = ''):
        self.output_dir = output_dir or os.path.join(os.getcwd(), 'reports')
        os.makedirs(self.output_dir, exist_ok=True)

    def export(self, report_data: Dict[str, Any], format: str = 'pdf',
               filename: str = '') -> str:
        """导出报告

        Args:
            report_data: 报告数据字典
            format: 'pdf' / 'markdown' / 'excel' / 'html'
            filename: 自定义文件名

        Returns:
            str: 导出文件路径
        """
        if format == 'pdf':
            return self._export_pdf(report_data, filename)
        elif format == 'markdown':
            return self._export_markdown(report_data, filename)
        elif format == 'excel':
            return self._export_excel(report_data, filename)
        elif format == 'html':
            return self._export_html(report_data, filename)
        else:
            raise ValueError(f"不支持的格式: {format}")

    def _export_markdown(self, data: Dict[str, Any], filename: str = '') -> str:
        """导出 Markdown 格式"""
        if not filename:
            filename = f"evaluation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"

        filepath = os.path.join(self.output_dir, filename)
        lines = self._build_markdown(data)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        logger.info(f"Markdown报告已导出: {filepath}")
        return filepath

    def _export_html(self, data: Dict[str, Any], filename: str = '') -> str:
        """导出 HTML 格式"""
        if not filename:
            filename = f"evaluation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"

        filepath = os.path.join(self.output_dir, filename)
        html = self._build_html(data)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html)

        logger.info(f"HTML报告已导出: {filepath}")
        return filepath

    def _export_pdf(self, data: Dict[str, Any], filename: str = '') -> str:
        """导出 PDF 格式"""
        if not filename:
            filename = f"evaluation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        filepath = os.path.join(self.output_dir, filename)

        try:
            # 尝试使用 WeasyPrint
            from weasyprint import HTML
            html = self._build_html(data)
            HTML(string=html).write_pdf(filepath)
        except ImportError:
            try:
                # 回退到 ReportLab
                filepath = self._export_pdf_reportlab(data, filename)
            except ImportError:
                # 最终回退: 保存HTML
                html_path = filepath.replace('.pdf', '.html')
                with open(html_path, 'w', encoding='utf-8') as f:
                    f.write(self._build_html(data))
                logger.warning("WeasyPrint/ReportLab 未安装, 已导出HTML格式")
                return html_path

        logger.info(f"PDF报告已导出: {filepath}")
        return filepath

    def _export_pdf_reportlab(self, data: Dict[str, Any], filename: str) -> str:
        """使用 ReportLab 导出 PDF"""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, PageBreak)
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        filepath = os.path.join(self.output_dir, filename)
        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                rightMargin=20*mm, leftMargin=20*mm,
                                topMargin=15*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        story = []

        # 标题
        title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                     fontSize=20, spaceAfter=20)
        story.append(Paragraph("座椅舒适性评测报告", title_style))
        story.append(Spacer(1, 10))

        # 基本信息
        story.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                               styles['Normal']))
        story.append(Spacer(1, 10))

        # 指标表格
        metrics = data.get('metrics', {})
        if metrics:
            story.append(Paragraph("指标汇总", styles['Heading2']))
            table_data = [['指标', '数值', '单位', '评级']]
            for mid, info in list(metrics.items())[:50]:
                table_data.append([
                    mid,
                    f"{info.get('value', 0):.3f}",
                    info.get('unit', ''),
                    info.get('grade', '-')
                ])

            if len(table_data) > 1:
                t = Table(table_data, colWidths=[120, 80, 60, 60])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16213e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
                ]))
                story.append(t)
                story.append(Spacer(1, 15))

        # 异常标注
        anomalies = data.get('anomalies', [])
        if anomalies:
            story.append(Paragraph("异常标注", styles['Heading2']))
            for a in anomalies[:20]:
                severity_color = {'critical': 'red', 'warning': 'orange', 'notice': 'yellow'}
                story.append(Paragraph(
                    f"[{a.get('severity', '?').upper()}] {a.get('metric_id', '?')}: "
                    f"{a.get('suggestion', '')}",
                    styles['Normal']
                ))
            story.append(Spacer(1, 10))

        # 诊断建议
        recommendations = data.get('recommendations', [])
        if recommendations:
            story.append(Paragraph("诊断建议", styles['Heading2']))
            for rec in recommendations:
                story.append(Paragraph(f"- {rec}", styles['Normal']))

        doc.build(story)
        return filepath

    def _export_excel(self, data: Dict[str, Any], filename: str = '') -> str:
        """导出 Excel 格式"""
        if not filename:
            filename = f"evaluation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()

            # Sheet 1: 指标汇总
            ws1 = wb.active
            ws1.title = "指标汇总"
            headers = ['指标ID', '数值', '单位', '位置', '组别', '质量', '评级']
            ws1.append(headers)

            header_fill = PatternFill(start_color='16213e', end_color='16213e', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            for col, h in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font

            metrics = data.get('metrics', {})
            for i, (mid, info) in enumerate(metrics.items(), 2):
                ws1.append([
                    mid,
                    info.get('value', 0),
                    info.get('unit', ''),
                    info.get('location', ''),
                    info.get('group', ''),
                    info.get('quality', 'normal'),
                    info.get('grade', '-')
                ])

            # 列宽自动调整
            for col in range(1, len(headers) + 1):
                ws1.column_dimensions[get_column_letter(col)].width = 15

            # Sheet 2: 异常标注
            anomalies = data.get('anomalies', [])
            if anomalies:
                ws2 = wb.create_sheet("异常标注")
                anomaly_headers = ['指标ID', '数值', '基线均值', '偏离σ', '严重度', '方向', '建议']
                ws2.append(anomaly_headers)
                for col, h in enumerate(anomaly_headers, 1):
                    cell = ws2.cell(row=1, column=col, value=h)
                    cell.fill = header_fill
                    cell.font = header_font

                for a in anomalies:
                    ws2.append([
                        a.get('metric_id', ''),
                        a.get('value', 0),
                        a.get('baseline_mean', 0),
                        a.get('deviation_sigma', 0),
                        a.get('severity', ''),
                        a.get('direction', ''),
                        a.get('suggestion', '')
                    ])

                # 严重度着色
                red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                orange_fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
                for row in range(2, len(anomalies) + 2):
                    severity = ws2.cell(row=row, column=5).value
                    if severity == 'critical':
                        for col in range(1, 8):
                            ws2.cell(row=row, column=col).fill = red_fill
                    elif severity == 'warning':
                        for col in range(1, 8):
                            ws2.cell(row=row, column=col).fill = orange_fill

                for col in range(1, 8):
                    ws2.column_dimensions[get_column_letter(col)].width = 18

            # Sheet 3: 基线对比
            comparison = data.get('comparison', {})
            if comparison:
                ws3 = wb.create_sheet("基线对比")
                comp_headers = ['指标ID', '基线值', '当前值', '变化%', '方向', '显著性']
                ws3.append(comp_headers)
                for col, h in enumerate(comp_headers, 1):
                    cell = ws3.cell(row=1, column=col, value=h)
                    cell.fill = header_fill
                    cell.font = header_font

                for mid, result in comparison.items():
                    ws3.append([
                        mid,
                        result.get('baseline_value', 0),
                        result.get('current_value', 0),
                        result.get('change_pct', 0),
                        result.get('direction', ''),
                        result.get('significance', '')
                    ])

                for col in range(1, 7):
                    ws3.column_dimensions[get_column_letter(col)].width = 18

            wb.save(filepath)
            logger.info(f"Excel报告已导出: {filepath}")

        except ImportError:
            # 回退到 CSV
            filepath = filepath.replace('.xlsx', '.csv')
            self._export_csv(data, filepath)
            logger.warning("openpyxl 未安装, 已导出CSV格式")

        return filepath

    def _export_csv(self, data: Dict[str, Any], filepath: str):
        """导出 CSV 格式"""
        import csv
        metrics = data.get('metrics', {})
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['指标ID', '数值', '单位', '位置', '组别', '质量'])
            for mid, info in metrics.items():
                writer.writerow([
                    mid,
                    info.get('value', 0),
                    info.get('unit', ''),
                    info.get('location', ''),
                    info.get('group', ''),
                    info.get('quality', 'normal')
                ])

    def _build_markdown(self, data: Dict[str, Any]) -> List[str]:
        """构建 Markdown 内容"""
        lines = []
        lines.append("# 座椅舒适性评测报告")
        lines.append("")
        lines.append(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")

        # 汇总信息
        summary = data.get('summary', {})
        if summary:
            lines.append("## 汇总信息")
            lines.append(f"- 总指标数: {summary.get('total_metrics', 0)}")
            lines.append(f"- 异常数: {summary.get('anomaly_count', 0)}")
            lines.append(f"- 风险等级: {summary.get('risk_level', 'unknown')}")
            lines.append("")

        # 指标表格
        metrics = data.get('metrics', {})
        if metrics:
            lines.append("## 指标汇总")
            lines.append("| 指标ID | 数值 | 单位 | 评级 |")
            lines.append("|:---|:---:|:---:|:---:|")
            for mid, info in list(metrics.items())[:50]:
                lines.append(
                    f"| {mid} | {info.get('value', 0):.3f} | "
                    f"{info.get('unit', '')} | {info.get('grade', '-')} |"
                )
            lines.append("")

        # 异常标注
        anomalies = data.get('anomalies', [])
        if anomalies:
            lines.append("## 异常标注")
            for a in anomalies:
                lines.append(
                    f"- **[{a.get('severity', '?').upper()}]** {a.get('metric_id', '?')}: "
                    f"{a.get('suggestion', '')} (偏离 {a.get('deviation_sigma', 0):.1f}σ)"
                )
            lines.append("")

        # 诊断建议
        recommendations = data.get('recommendations', [])
        if recommendations:
            lines.append("## 诊断建议")
            for rec in recommendations:
                lines.append(f"- {rec}")
            lines.append("")

        return lines

    def _build_html(self, data: Dict[str, Any]) -> str:
        """构建 HTML 内容"""
        metrics = data.get('metrics', {})
        anomalies = data.get('anomalies', [])
        recommendations = data.get('recommendations', [])
        summary = data.get('summary', {})

        # 指标表格行
        metric_rows = ''
        for mid, info in list(metrics.items())[:50]:
            metric_rows += f'''
            <tr>
                <td>{mid}</td>
                <td>{info.get('value', 0):.3f}</td>
                <td>{info.get('unit', '')}</td>
                <td>{info.get('location', '')}</td>
                <td>{info.get('grade', '-')}</td>
            </tr>'''

        anomaly_rows = ''
        for a in anomalies:
            anomaly_rows += f'''
            <tr class="severity-{a.get('severity', 'notice')}">
                <td>{a.get('metric_id', '')}</td>
                <td>{a.get('value', 0):.3f}</td>
                <td>{a.get('deviation_sigma', 0):.1f}σ</td>
                <td>{a.get('severity', '')}</td>
                <td>{a.get('suggestion', '')}</td>
            </tr>'''

        rec_items = ''
        for rec in recommendations:
            rec_items += f'<li>{rec}</li>'

        return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>座椅舒适性评测报告</title>
<style>
    body {{ font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif; 
           background: #1a1a2e; color: #e0e0e0; margin: 20px; }}
    h1 {{ color: #2196F3; }}
    h2 {{ color: #4CAF50; border-bottom: 1px solid #333; padding-bottom: 5px; }}
    table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
    th {{ background: #16213e; color: #aaa; padding: 8px; text-align: left; }}
    td {{ padding: 6px 8px; border-bottom: 1px solid #333; }}
    tr:hover {{ background: #2a2a4e; }}
    .severity-critical {{ background: #4a1a1a; }}
    .severity-warning {{ background: #3a3a1a; }}
    .summary-box {{ background: #16213e; border: 1px solid #333; border-radius: 8px; padding: 15px; margin: 10px 0; }}
    .risk-critical {{ color: #F44336; }}
    .risk-high {{ color: #FF9800; }}
    .risk-medium {{ color: #FFC107; }}
    .risk-low {{ color: #8BC34A; }}
    .risk-normal {{ color: #4CAF50; }}
</style>
</head>
<body>
<h1>座椅舒适性评测报告</h1>
<p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>

<div class="summary-box">
    <h2>汇总信息</h2>
    <p>总指标数: <strong>{summary.get('total_metrics', 0)}</strong> | 
       异常数: <strong>{summary.get('anomaly_count', 0)}</strong> | 
       风险等级: <strong class="risk-{summary.get('risk_level', 'normal')}">{summary.get('risk_level', 'normal')}</strong></p>
</div>

<h2>指标汇总</h2>
<table>
    <tr><th>指标ID</th><th>数值</th><th>单位</th><th>位置</th><th>评级</th></tr>
    {metric_rows}
</table>

<h2>异常标注</h2>
<table>
    <tr><th>指标ID</th><th>数值</th><th>偏离</th><th>严重度</th><th>建议</th></tr>
    {anomaly_rows}
</table>

<h2>诊断建议</h2>
<ul>{rec_items}</ul>

<p style="color: #666; font-size: 11px; margin-top: 30px;">
    报告由全量统计分析模块自动生成 | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
</p>
</body>
</html>'''